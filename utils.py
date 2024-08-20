#!/usr/bin/env python3
# for testing local models responses on the droplet

"""
Module Docstring
"""

__author__ = "Alex Bishop"
__version__ = "0.3.0"
__license__ = "MIT"

# utils.py
import os
import pandas as pd
import logging
from generation import generate
from config import *
from user_messages import *
import uuid
import json
from openpyxl import Workbook, load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path

def multi_selection_input(prompt, items):
    while True:
        print(prompt)
        for idx, item in enumerate(items, start=1):
            print(f"{idx}. {PROMPT_COLOR}{item}{RESET_STYLE}")
        selection_input = input(msg_enter_prompt_selection_multiple()).strip()

        if not selection_input:  # Handle empty input
            print(msg_enter_prompt_selection_multiple())
            continue

        selected_indices = []
        for part in selection_input.split(','):
            try:
                if '-' in part:
                    start, end = map(int, part.split('-'))
                    selected_indices.extend(range(start, end + 1))
                else:
                    selected_indices.append(int(part))
            except ValueError:
                print(msg_invalid_number())
                break  # Break out of the for loop, continue while loop for new input
        else:  # This else corresponds to the for loop
            # Deduplicate and sort the indices
            selected_indices = sorted(set(selected_indices))

            # Validate selection
            try:
                selected_items = [items[idx - 1] for idx in selected_indices]
                print(msg_prompt_confirm_multi())
                for item in selected_items:
                    print(msg_list_selected_prompts(item))
                if confirm_selection():
                    return selected_items
            except IndexError:
                print(msg_invalid_retry())

def confirm_selection(message=msg_confirm_selection()):
    while True:
        confirm = input(message).strip().lower()
        # Treat empty input as 'yes'
        if confirm in ['y', 'yes', '']:
            return True
        elif confirm in ['n', 'no']:
            return False
        else:
            print(msg_select_y_n())

def select_category(categories):
    print(msg_select_category())
    for idx, category in enumerate(categories):
        print(f"{idx + 1}. {CATEGORY_COLOR}{category}{RESET_STYLE}")
    print(msg_custom_prompt_instr())
    print(msg_q_to_quit())

    while True:
        category_input = input(msg_enter_category_num()).strip()
        if category_input.lower() == 'q':
            return None
        elif category_input == '':
            print(msg_enter_category_num())
            continue

        try:
            category_idx = int(category_input) - 1
            if category_idx == -1:
                selected_category = 'custom'
            elif 0 <= category_idx < len(categories):
                selected_category = categories[category_idx]
                if not confirm_selection(msg_confirm_custom_cat(selected_category)):
                    print(msg_invalid_retry())
                    continue  # Stay in the loop for a new selection
            else:
                raise ValueError  # Treat out-of-range numbers as invalid
        except ValueError:
            print(msg_invalid_number())
            continue  # Stay in the loop for a new selection

        return selected_category

def print_response_stats(response, response_time, char_count, word_count):
    if response_time > 60:
        minutes = int(response_time // 60)
        seconds = int(response_time % 60)
        formatted_response_time = f"{minutes} minutes and {seconds} seconds"
    else:
        formatted_response_time = f"{response_time:.2f} seconds"

    print(f"\n{BOLD_EFFECT}{STATS_COLOR}Response Time:{RESET_STYLE} {BOLD_EFFECT}{formatted_response_time}{RESET_STYLE}")
    print(f"{BOLD_EFFECT}{STATS_COLOR}Character Count:{RESET_STYLE} {BOLD_EFFECT}{char_count}{RESET_STYLE}")
    print(f"{BOLD_EFFECT}{STATS_COLOR}Word Count:{RESET_STYLE} {BOLD_EFFECT}{word_count}{RESET_STYLE}")
    character_rate = char_count / response_time if response_time > 0 else 0
    word_rate = word_count / response_time if response_time > 0 else 0
    print(f"{BOLD_EFFECT}{STATS_COLOR}Character Rate:{RESET_STYLE} {BOLD_EFFECT}{character_rate:.2f}{RESET_STYLE} characters per second")
    print(f"{BOLD_EFFECT}{STATS_COLOR}Word Rate:{RESET_STYLE} {BOLD_EFFECT}{word_rate:.2f}{RESET_STYLE} words per second")

def get_user_rating():
    while True:
        rating = input(msg_get_response_rating()).strip()
        try:
            rating = int(rating)
            if 1 <= rating <= 5:
                return rating
            else:
                print(msg_invalid_rating_num())
        except ValueError:
            print(msg_invalid_number())

def process_excel_file(model_name, prompt, excel_path):
    # Load the Excel file
    df = pd.read_excel(excel_path, engine=excel_engine)

    # Define the new column name based on the model name
    summary_column_name = f"{model_name}-Summary"

    # Ensure the new summary column exists, if not, create it
    if summary_column_name not in df.columns:
        df[summary_column_name] = pd.Series(dtype='object')

    # Iterate through each row in the DataFrame
    for index, row in df.iterrows():
        content = row['Message_Content']  # Assuming the content is in column B
        # Extract the first 15 words from the content
        first_15_words = ' '.join(content.split()[:summary_excerpt_wordcount])
        
        # Print the message including the first 15 words of the content
        print(msg_generating_msg(model_name, prompt))
        print(f"'{first_15_words}...'")
        
        # Generate the summary using the selected model and prompt
        try:
            _, response, _, _, _ = generate(model_name, f"{PROMPT_COLOR}{prompt}{RESET_STYLE} {content}", None)
        except Exception as e:
            logging.error(msg_error_response(prompt, e))
            response = msg_error_simple(e)
        # Prepend the prompt to the response
        full_response = f"{prompt}\n{response}"
        df.at[index, summary_column_name] = full_response  # Write the prompt and response to the new summary column
    
    # Save the modified DataFrame back to the Excel file
    df.to_excel(excel_path, index=False, engine=excel_engine)
    print(msg_excel_completed(excel_path))

def export_json_prompts():
    print("Exporting all prompts to Excel...")
    
    # Define the Excel file path
    excel_path = 'prompts_export.xlsx'
    
    # Load prompts from JSON file
    with open('prompts.json', 'r') as file:
        prompts_data = json.load(file)
    
    # Prepare data for DataFrame
    data = []
    for category, prompts in prompts_data['categories'].items():
        for prompt in prompts:
            # Generate a UUID for each prompt
            prompt_id = str(uuid.uuid4())
            data.append({
                'Prompt_ID': prompt_id,
                'Prompt_Category': category,
                'Prompt_Text': prompt
            })
    
    # Create a DataFrame
    df = pd.DataFrame(data)
    
    # Check if the Excel file exists
    if Path(excel_path).exists():
        print(f"File {excel_path} exists, it will be overwritten.")
    
    # Create a new Excel workbook and sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "Prompts"
    
    # Add DataFrame to Excel
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Save the workbook
    wb.save(filename=excel_path)
    print("Prompts exported successfully.")

def list_response_files():
    response_dir = 'responses'  # Adjust path as necessary
    files = [f for f in os.listdir(response_dir) if f.endswith('.json')]
    files.sort()  # Alphabetize the list
    return files

def select_response_files():
    files = list_response_files()
    print(f"{msg_word_select()} the {msg_word_model()} response file(s) to export:")
    selected_files = multi_selection_input("Enter the numbers (e.g., 1,2,3): ", files)
    return selected_files

def process_json_files(files):
    prompts_df = pd.read_excel('prompts.xlsx', engine='openpyxl')
    data = []

    for file in files:
        with open(f'responses/{file}', 'r') as f:
            responses = json.load(f)
            for response in responses:
                prompt_text = response['prompt']
                prompt_uuid = prompts_df[prompts_df['Prompt_Text'] == prompt_text]['Prompt_ID'].values[0] if not prompts_df[prompts_df['Prompt_Text'] == prompt_text].empty else ''
                data.append({
                    'Message_ID': str(uuid.uuid4()),
                    'Conv_ID': f"test-{file.replace('.json', '')}",
                    'Prompt_ID': prompt_uuid,
                    'Prompt_Text': prompt_text,
                    'Prompt_Category': '',
                    'Input_Text': '',
                    'Benchmark_Response': '',
                    'Overall_Rating': '',
                    'User_Rating': response.get('rating', ''),
                    'Clarity': '',
                    'Accuracy': '',
                    'Conciseness': '',
                    'Response_Dur': response['response_time'],
                    'Msg_Timestamp': '',
                    'Msg_Month': '',
                    'Msg_Year': '',
                    'Msg_AuthorRole': 'assistant',
                    'Msg_AuthorName': '',
                    'GPT_Name': file.replace('.json', ''),
                    'GPT_ID': '',
                    'Sequence_Number': '',
                    'Msg_Content': response['response'],
                    'Stored_Memory': '',
                    'Msg_Status': 'exported',
                    'Msg_EndTurn': '',
                    'Msg_Weight': '',
                    'Msg_VoiceMode': '',
                    'Msg_Metadata': '',
                    'Msg_Parent_ID': '',
                    'Msg_Children_IDs': '',
                    'Sentiment_Polarity': '',
                    'Sentiment_Subjectivity': '',
                    'URL_List': '',
                    'Code_Related': '',
                    'Chars_Total': response['char_count'],
                    'Sentences_Total': '',
                    'Words_Total': response['word_count'],
                    'Tokens_Total': '',
                    'Img_Generated': '',
                    'Img_URL': '',
                    'Img_Size_Bytes': '',
                    'Img_Width': '',
                    'Img_Height': '',
                    'Img_Fovea': '',
                    'Img_Gen_ID': '',
                    'Img_Prompt': '',
                    'Img_Seed': ''
                })
    return pd.DataFrame(data)

def export_to_excel(df):
    excel_path = 'prompt_responses.xlsx'
    # Check if the Excel file exists to determine if we need to append or write new
    if os.path.exists(excel_path):
        # Load the existing workbook
        book = load_workbook(excel_path)
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            writer.book = book
            # If the 'Responses' sheet exists, find the next available row
            if 'Responses' in writer.book.sheetnames:
                startrow = writer.book['Responses'].max_row
            else:
                startrow = 0
            # Write the dataframe to the 'Responses' sheet in the existing workbook
            df.to_excel(writer, index=False, sheet_name='Responses', startrow=startrow)
            # No need to call save() as it's handled by the context manager
    else:
        # If the file doesn't exist, simply write the dataframe to a new file
        with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Responses')
            # No need to call save() as it's handled by the context manager

def export_all_responses():
    selected_files = select_response_files()
    if not selected_files:
        print("No files selected.")
        return
    
    if confirm_selection():
        df = process_json_files(selected_files)
        export_to_excel(df)
        print("Export completed successfully.")